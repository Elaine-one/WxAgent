import openpyxl


def read_excel(path: str) -> dict[str, list[list]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        result[name] = [[str(cell.value) if cell.value is not None else ""
                         for cell in row] for row in ws.iter_rows()]
    return result
